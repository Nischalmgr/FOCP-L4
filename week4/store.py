import tkinter as tk
from tkinter import ttk
import openpyxl 
from openpyxl import Workbook
File_name= "classwork.xlsx"
def inttilizeExcel():
    try:
        workbook= openpyxl.load_workbook(File_name)    
        worksheet=workbook.active
    except FileNotFoundError:
        workbook = Workbook()

        worksheet=workbook.active
        worksheet.append(["Name", "marks"])
        workbook.save(File_name)
 

        
        
       

root=tk.Tk()
root.title("student marks")
root.geometry("500x500")
frame=tk.Frame(root)
frame.grid(row=0, column=0)
tk.Label(frame, text="name").grid(row=0 , column=0)
name=tk.Entry(frame)
name.grid(row=0, column=1,columnspan=1)
tk.Label(frame,text="marks").grid(row=1,column=0)
marks=tk.Entry(frame)
marks.grid(row=1, column=1)

def add(name,value):
    tree.insert("",tk.END, values=(name, value))

def insert_exel():
    name_=name.get()
    marks_=marks.get()
    workbook=openpyxl.load_workbook(File_name)
    worksheet=workbook.active
    worksheet.append((name_,marks_))
    workbook.save(File_name)
    add(name_,marks_)




tree=ttk.Treeview(frame, columns=("Name", "marks") ,show='headings')
tree.heading("Name", text="Name")
tree.column("Name", anchor=tk.CENTER)
tree.heading("marks", text="Mark")
tree.grid(row=4, columnspan=2)


inttilizeExcel()
tk.Button(frame, text="ok" ,command=insert_exel).grid(row=2 , column=0)
 
root.mainloop()